import pandas as pd
import io
import pypdf
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font

def get_active_classes_from_timetable(timetable_file, exam_date, exam_session):
    if not timetable_file:
        return []
    
    filename = timetable_file.name.lower()
    
    # Generate Date Variations for robust matching
    date_variations = []
    import re
    try:
        import pandas as pd
        dt = pd.to_datetime(exam_date, dayfirst=True)
        date_variations.append(dt.strftime('%d%m%Y')) # 09042026
        date_variations.append(f"{dt.day}{dt.month}{dt.year}") # 942026
        date_variations.append(f"{dt.day:02d}{dt.month}{dt.year}") # 0942026
        date_variations.append(f"{dt.day}{dt.month:02d}{dt.year}") # 9042026
        date_variations.append(dt.strftime('%d%b%Y').upper()) # 09APR2026
        date_variations.append(dt.strftime('%d%B%Y').upper()) # 09APRIL2026
        date_variations.append(f"{dt.day}{dt.strftime('%b').upper()}{dt.year}") # 9APR2026
        date_variations.append(f"{dt.day}{dt.strftime('%B').upper()}{dt.year}") # 9APRIL2026
    except:
        date_variations.append(re.sub(r'\D', '', exam_date))
        
    clean_session = exam_session.upper().strip()
    active_classes = set()
    
    # Generic Tracker Logic
    def parse_line(line, recent_dept, recent_year):
        upper_line = line.upper()
        
        # Create a heavily cleaned version for department matching (no commas, dots, slashes)
        # We pad it with spaces to ensure word boundaries match
        clean_for_match = f" {upper_line.replace(',', ' ').replace('.', ' ').replace('-', ' ')} "
        
        temp_dept = []
        if 'COMPUTER SCIENCE' in upper_line or ' CS ' in clean_for_match: temp_dept.append('CS')
        if 'ARTIFICIAL INTELLIGENCE' in upper_line or 'DATA SCIENCE' in upper_line or 'AI&DS' in upper_line or 'AI DS' in clean_for_match or ' AIDS ' in clean_for_match: temp_dept.append('AIDS')
        if 'COMMERCE' in upper_line or ' BCOM ' in clean_for_match or ' COM ' in clean_for_match: temp_dept.append('BCOM')
        if 'TAMIL' in upper_line or ' T ' in clean_for_match or ' TA ' in clean_for_match: temp_dept.append('B A TAMIL')
        if 'CHEMISTRY' in upper_line or ' CHE ' in clean_for_match or ' CH ' in clean_for_match: temp_dept.append('CHE')
        if 'COMPUTER APPLICATIONS' in upper_line or ' BCA ' in clean_for_match: temp_dept.append('BCA')
        if 'BUSINESS ADMINISTRATION' in upper_line or ' BBA ' in clean_for_match: temp_dept.append('BBA')
        if ' MIB ' in clean_for_match: temp_dept.append('MIB')
        if 'CORPORATE' in upper_line or ' CCA ' in clean_for_match: temp_dept.append('CCA')
        if 'BCOM CA' in upper_line or 'B.COM CA' in upper_line or 'BOM CA' in upper_line or ' BCOM CA ' in clean_for_match: temp_dept.append('BCOM CA')
        
        if temp_dept:
            recent_dept = temp_dept
            
        yrs = []
        clean_upper = upper_line.replace(' ', '')
        exact_match = re.search(r'EXACTYEAR=([I/]+)', clean_upper)
        
        target_str = exact_match.group(1) if exact_match else ""
        
        if target_str in ['I/II/III', 'III/II/I'] or (not exact_match and ('I/II/III' in clean_upper or 'III/II/I' in clean_upper)):
            yrs = ['I-YEAR', 'II-YEAR', 'III-YEAR']
        elif target_str in ['II/III', 'III/II'] or (not exact_match and ('II/III' in clean_upper or 'III/II' in clean_upper)):
            yrs = ['II-YEAR', 'III-YEAR']
        elif target_str in ['I/II', 'II/I'] or (not exact_match and ('I/II' in clean_upper or 'II/I' in clean_upper)):
            yrs = ['I-YEAR', 'II-YEAR']
        elif target_str == 'I':
            yrs = ['I-YEAR']
        elif target_str == 'II':
            yrs = ['II-YEAR']
        elif target_str == 'III':
            yrs = ['III-YEAR']
        else:
            yr = extract_year_from_text(upper_line)
            if not yr and 'SEMESTER' in upper_line:
                match = re.search(r'SEMESTER\s*[:\-]?\s*(I{1,3}|IV|V|VI|[1-6])', upper_line)
                if match:
                    sem_str = match.group(1)
                    if sem_str in ['1', '2', 'I', 'II']: yr = 'I-YEAR'
                    elif sem_str in ['3', '4', 'III', 'IV']: yr = 'II-YEAR'
                    elif sem_str in ['5', '6', 'V', 'VI']: yr = 'III-YEAR'
            if yr:
                yrs = [yr]
                
        if yrs:
            recent_year = yrs
            
        clean_line = re.sub(r'[\s\-/\.]', '', upper_line)
        
        # Check if ANY of our date variations are in the line
        date_matched = False
        for dv in date_variations:
            if dv and dv in clean_line:
                date_matched = True
                break
                
        session_matched = False
        if clean_session == 'FN' and ('FN' in clean_line or 'FORENOON' in clean_line or 'MORNING' in clean_line):
            session_matched = True
        elif clean_session == 'AN' and ('AN' in clean_line or 'AFTERNOON' in clean_line or 'EVENING' in clean_line):
            session_matched = True

        if date_matched and session_matched:
            matched_depts = temp_dept if temp_dept else recent_dept
            matched_years = yrs if yrs else (recent_year if isinstance(recent_year, list) else [recent_year] if recent_year else [])
            
            if 'FOUNDATION' in upper_line or 'ALL' in upper_line.split():
                matched_depts = ['CS', 'AIDS', 'BCOM', 'B A TAMIL', 'CHE', 'BCA', 'BBA', 'MIB', 'CCA']
                
            for d in matched_depts:
                for y in matched_years:
                    if y:
                        active_classes.add(f"{d} - {y}")
                    
        return recent_dept, recent_year

    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        if hasattr(timetable_file, 'seek'):
            timetable_file.seek(0)
        df = pd.read_excel(timetable_file, header=None)
        recent_dept = []
        recent_year = ''
        
        current_date_str = ''
        current_session = ''
        
        for idx, row in df.iterrows():
            row_vals = []
            exact_year_str = ""
            for val in row.values:
                if pd.isna(val):
                    continue
                sval = str(val).strip().upper()
                # Check for Date
                if re.search(r'\d{2}-\d{2}-\d{4}', sval) or re.search(r'\d{4}-\d{2}-\d{2}', sval):
                    current_date_str = sval
                elif isinstance(val, pd.Timestamp):
                    current_date_str = val.strftime('%d-%m-%Y')
                # Check for Session
                if sval in ['FN', 'AN', 'FORENOON', 'AFTERNOON']:
                    if sval == 'FORENOON': current_session = 'FN'
                    elif sval == 'AFTERNOON': current_session = 'AN'
                    else: current_session = sval
                    
                # Check for standalone exact year cell
                c_up = sval.replace(' ', '')
                if c_up in ['I', 'II', 'III', 'I/II', 'II/I', 'II/III', 'III/II', 'I/II/III', 'III/II/I', 'I/III', 'III/I']:
                    exact_year_str = c_up
                    
                row_vals.append(sval)
                
            if not row_vals: continue
            
            row_str = ' '.join(row_vals)
            full_row_str = f"{row_str} {current_date_str} {current_session}"
            if exact_year_str:
                full_row_str = f"EXACTYEAR={exact_year_str} {full_row_str}"
            
            recent_dept, recent_year = parse_line(full_row_str, recent_dept, recent_year)
    else:
        if hasattr(timetable_file, 'seek'):
            timetable_file.seek(0)
        import pypdf
        reader = pypdf.PdfReader(timetable_file)
        recent_dept = []
        recent_year = ''
        for page in reader.pages:
            text = page.extract_text()
            if not text:
                continue
            for line in text.split('\n'):
                recent_dept, recent_year = parse_line(line, recent_dept, recent_year)

    return sorted(list(set(active_classes)))

import re

def extract_year_from_text(text):
    if not text:
        return None
    text = str(text).upper()
    # Aggressively look for I, II, III, 1, 2, 3 followed by YEAR, YR, or degree names
    pattern = r'\b(I{1,3}|1ST|2ND|3RD|[1-3])\s*[-_]?\s*(YEAR|YR|B\.?SC|B\.?COM|BCA|BBA|B\.?A\b|DEGREE|BRANCH|SEM)'
    match = re.search(pattern, text)
    if match:
        val = match.group(1).replace('ST','').replace('ND','').replace('RD','')
        if val == 'I' or val == '1': return 'I-YEAR'
        if val == 'II' or val == '2': return 'II-YEAR'
        if val == 'III' or val == '3': return 'III-YEAR'
    
    # If still not found, check if the text contains EXACTLY something like "I YEAR" somewhere else
    if '1ST YEAR' in text or 'I YEAR' in text or 'I-YEAR' in text: return 'I-YEAR'
    if '2ND YEAR' in text or 'II YEAR' in text or 'II-YEAR' in text: return 'II-YEAR'
    if '3RD YEAR' in text or 'III YEAR' in text or 'III-YEAR' in text: return 'III-YEAR'
    
    return None

def extract_student_data(df_sheet, fallback_year=None):
    header_idx = -1
    reg_col = None
    top_level_year = fallback_year
    
    # 1. Scan for YEAR in the top rows (0 to 20)
    for i in range(min(20, len(df_sheet))):
        row_vals = [str(x).strip().upper() for x in df_sheet.iloc[i].tolist() if pd.notna(x)]
        full_row = " ".join(row_vals)
        yr = extract_year_from_text(full_row)
        if yr and not top_level_year:
            top_level_year = yr

    # 2. Find the header row by looking for REG/ROLL/REGISTER
    for i in range(min(25, len(df_sheet))):
        row_vals = [str(x).strip().upper() for x in df_sheet.iloc[i].tolist()]
        if any('REG' in val or 'ROLL' in val or 'REGISTER' in val for val in row_vals):
            header_idx = i
            break

    if header_idx != -1:
        headers = df_sheet.iloc[header_idx].tolist()
        df_data = df_sheet.iloc[header_idx+1:].copy()
        # Ensure column names are unique strings
        df_data.columns = [str(c) if pd.notna(c) else f"Unnamed_{j}" for j, c in enumerate(headers)]
        
        # 3. Explicit Header Check (Highest Priority)
        for c in df_data.columns:
            cup = str(c).upper()
            if 'SUB REG NO' in cup or 'SUB REG' in cup:
                reg_col = c
                break
                
        if not reg_col:
            for c in df_data.columns:
                cup = str(c).upper()
                if ('REG' in cup or 'ROLL' in cup or 'ID' in cup) and 'NAME' not in cup:
                    reg_col = c
                    break
    else:
        df_data = df_sheet.copy()
        
    # 4. Find reg_col purely by analyzing column contents (score-based fallback)
    if not reg_col:
        best_col = None
        best_score = -1
        
        for c in df_data.columns:
            col_data = df_data[c].dropna().astype(str)
            if len(col_data) == 0:
                continue
                
            # Count how many cells look like a register number (>=2 chars, contains digit)
            valid_count = sum(1 for x in col_data if len(x.strip()) >= 2 and any(char.isdigit() for char in x))
            score = valid_count / len(col_data)
            
            # Penalize the S.No column implicitly if it's strictly sequential 1,2,3
            if valid_count > 0 and 'NAME' not in str(c).upper():
                if score > best_score:
                    best_score = score
                    best_col = c
                
        if best_score > 0:
            reg_col = best_col
        else:
            reg_col = df_data.columns[0] if len(df_data.columns) > 0 else None

    # 4. Check for specific YEAR column
    year_col = None
    for c in df_data.columns:
        if isinstance(c, str) and 'YEAR' in c.upper():
            year_col = c
            break
            
    return df_data, reg_col, year_col, top_level_year

def get_student_classes(student_files):
    classes = set()
    for s_file in student_files:
        file_yr = extract_year_from_text(s_file.name)
        if s_file.name.lower().endswith('.pdf'):
            import pypdf
            reader = pypdf.PdfReader(s_file)
            top_level_year = file_yr
            sheet = s_file.name.replace('.pdf', '')
            if not top_level_year:
                for page in reader.pages:
                    text = page.extract_text()
                    if text:
                        yr = extract_year_from_text(text)
                        if yr:
                            top_level_year = yr
                            break
            if not top_level_year:
                top_level_year = "UNKNOWN YEAR"
            classes.add(f"{sheet} - {top_level_year}")
        elif s_file.name.lower().endswith('.csv'):
            if hasattr(s_file, 'seek'): s_file.seek(0)
            sheet = s_file.name.replace('.csv', '')
            df_sheet = pd.read_csv(s_file)
            if len(df_sheet) > 0:
                sheet_yr = extract_year_from_text(sheet)
                fallback = sheet_yr if sheet_yr else file_yr
                df_data, reg_col, year_col, top_level_year = extract_student_data(df_sheet, fallback_year=fallback)

                if year_col:
                    unique_years = df_data[year_col].dropna().unique()
                    for y in unique_years:
                        classes.add(f"{sheet} - {str(y).strip()}")
                elif top_level_year:
                    classes.add(f"{sheet} - {top_level_year}")
                else:
                    classes.add(f"{sheet} - UNKNOWN YEAR")
        else:
            xls_students = pd.ExcelFile(s_file)
            for sheet in xls_students.sheet_names:
                df_sheet = pd.read_excel(xls_students, sheet_name=sheet)
                if len(df_sheet) > 0:
                    sheet_yr = extract_year_from_text(sheet)
                    fallback = sheet_yr if sheet_yr else file_yr
                    df_data, reg_col, year_col, top_level_year = extract_student_data(df_sheet, fallback_year=fallback)
                    
                    if year_col:
                        unique_years = df_data[year_col].dropna().unique()
                        for y in unique_years:
                            classes.add(f"{sheet} - {str(y).strip()}")
                    elif top_level_year:
                        classes.add(f"{sheet} - {top_level_year}")
                    else:
                        classes.add(f"{sheet} - UNKNOWN YEAR")
    return sorted(list(classes))

def process_allocation(hall_file, student_files, timetable_file, exam_date, exam_session, selected_classes, exam_title):
    # 1. Read Hall Details
    if hall_file.name.lower().endswith('.pdf'):
        import pypdf
        reader = pypdf.PdfReader(hall_file)
        halls = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                for line in text.split('\n'):
                    parts = [p.strip() for p in line.split() if p.strip()]
                    if len(parts) >= 2:
                        hall_no = parts[0]
                        try:
                            total_seat = int(parts[1])
                            if len(parts) >= 4:
                                try:
                                    rows = int(parts[2])
                                    cols = int(parts[3])
                                    if 0 < total_seat <= 200:
                                        halls.append({'HALL NO': hall_no, 'TOTAL SEAT': total_seat, 'ROWS': rows, 'COLUMNS': cols})
                                    continue
                                except:
                                    pass
                            
                            # If no rows/cols provided or parsing failed, just store TOTAL SEAT
                            if 0 < total_seat <= 200:
                                halls.append({'HALL NO': hall_no, 'TOTAL SEAT': total_seat})
                        except:
                            pass
        if not halls:
            raise ValueError("Could not find valid hall data in the PDF. Ensure lines look like: '101 30 5 6'")
        df_halls = pd.DataFrame(halls)
    elif hall_file.name.lower().endswith('.csv'):
        df_halls = pd.read_csv(hall_file)
    else:
        df_halls = pd.read_excel(hall_file)
        
    # Standardize column names for halls to be extremely flexible
    df_halls.columns = df_halls.columns.astype(str).str.strip().str.upper()
    col_mapping = {}
    import re
    for col in df_halls.columns:
        if re.search(r'HALL|ROOM|CLASS', col):
            if 'HALL NO' not in col_mapping.values():
                col_mapping[col] = 'HALL NO'
        elif re.search(r'TOTAL|CAPACITY|SEAT|COUNT', col):
            if 'TOTAL SEAT' not in col_mapping.values():
                col_mapping[col] = 'TOTAL SEAT'
        elif re.search(r'ROW', col):
            col_mapping[col] = 'ROWS'
        elif re.search(r'COL', col):
            col_mapping[col] = 'COLUMNS'
            
    df_halls.rename(columns=col_mapping, inplace=True)
    
    if 'HALL NO' not in df_halls.columns and len(df_halls.columns) > 0:
        df_halls.rename(columns={df_halls.columns[0]: 'HALL NO'}, inplace=True)
    
    if 'TOTAL SEAT' not in df_halls.columns:
        if len(df_halls.columns) > 1:
            # Assume second column is total seats if not explicitly found
            df_halls.rename(columns={df_halls.columns[1]: 'TOTAL SEAT'}, inplace=True)
        elif len(df_halls.columns) == 1:
            # If they only provided 1 column, assume it is TOTAL SEAT
            df_halls['TOTAL SEAT'] = df_halls['HALL NO']
            # And assign dummy hall numbers
            df_halls['HALL NO'] = [f"H-{i+1}" for i in range(len(df_halls))]

    # 2. Read Student Details (read all sheets from all files)
    all_students = []
    
    for s_file in student_files:
        file_yr = extract_year_from_text(s_file.name)
        if s_file.name.lower().endswith('.pdf'):
            import pypdf
            reader = pypdf.PdfReader(s_file)
            sheet = s_file.name.replace('.pdf', '')
            top_level_year = file_yr
            
            # First pass to find year if not in filename
            if not top_level_year:
                for page in reader.pages:
                    text = page.extract_text()
                    if text:
                        yr = extract_year_from_text(text)
                        if yr:
                            top_level_year = yr
                            break
            if not top_level_year:
                top_level_year = "UNKNOWN YEAR"
                        
            class_key = f"{sheet} - {top_level_year}"
            
            # Second pass to extract reg numbers
            if class_key in selected_classes:
                for page in reader.pages:
                    text = page.extract_text()
                    if text:
                        words = text.split()
                        for w in words:
                            w = w.strip()
                            # Heuristic: looks like a reg number (>= 5 chars, has digit, and isn't purely punctuation)
                            if len(w) >= 3 and any(c.isdigit() for c in w) and any(c.isalnum() for c in w):
                                import re
                                matches = re.findall(r'[A-Za-z0-9]+', w)
                                best_match = w
                                for m in matches:
                                    if any(ch.isdigit() for ch in m):
                                        if len(m) >= 2:
                                            best_match = m
                                            break
                                        
                                all_students.append({
                                    'REG_NO': best_match,
                                    'NAME': '', # Hard to extract names reliably from PDF
                                    'DEPT': class_key
                                })
        elif s_file.name.lower().endswith('.csv'):
            if hasattr(s_file, 'seek'): s_file.seek(0)
            sheet = s_file.name.replace('.csv', '')
            df_sheet = pd.read_csv(s_file)
            if len(df_sheet) > 0:
                sheet_yr = extract_year_from_text(sheet)
                fallback = sheet_yr if sheet_yr else file_yr
                df_data, reg_col, year_col, top_level_year = extract_student_data(df_sheet, fallback_year=fallback)

                if reg_col:
                    # Clean up empty rows based on reg_col
                    df_data = df_data.dropna(subset=[reg_col])

                    for _, row in df_data.iterrows():
                        # Determine the year for this specific student
                        if year_col and not pd.isna(row[year_col]):
                            student_year = str(row[year_col]).strip()
                        elif top_level_year:
                            student_year = top_level_year
                        else:
                            student_year = "UNKNOWN YEAR"

                        class_key = f"{sheet} - {student_year}"
                        
                        if class_key in selected_classes:
                            name_val = ''
                            cols_list = list(df_data.columns)
                            for c in cols_list:
                                if isinstance(c, str) and 'NAME' in c.upper():
                                    name_val = row[c]
                                    break
                                    
                            raw_reg = str(row[reg_col])
                            import re
                            matches = re.findall(r'[A-Za-z0-9]+', raw_reg)
                            best_match = raw_reg
                            for m in matches:
                                if any(ch.isdigit() for ch in m):
                                    if len(m) >= 2:
                                        best_match = m
                                        break
                                    
                            all_students.append({
                                'REG_NO': best_match,
                                'NAME': name_val,
                                'DEPT': class_key
                            })
        else:
            xls_students = pd.ExcelFile(s_file)
            for sheet in xls_students.sheet_names:
                df_sheet = pd.read_excel(xls_students, sheet_name=sheet)
                if len(df_sheet) > 0:
                    sheet_yr = extract_year_from_text(sheet)
                    fallback = sheet_yr if sheet_yr else file_yr
                    df_data, reg_col, year_col, top_level_year = extract_student_data(df_sheet, fallback_year=fallback)
                    
                    if reg_col:
                        # Clean up empty rows based on reg_col
                        df_data = df_data.dropna(subset=[reg_col])
                        
                        for _, row in df_data.iterrows():
                            # Determine the year for this specific student
                            if year_col and not pd.isna(row[year_col]):
                                student_year = str(row[year_col]).strip()
                            elif top_level_year:
                                student_year = top_level_year
                            else:
                                student_year = "UNKNOWN YEAR"
                                
                            class_key = f"{sheet} - {student_year}"
                            
                            if class_key in selected_classes:
                                name_val = ''
                                cols_list = list(df_data.columns)
                                for c in cols_list:
                                    if isinstance(c, str) and 'NAME' in c.upper():
                                        name_val = row[c]
                                        break
                                        
                                raw_reg = str(row[reg_col])
                                import re
                                matches = re.findall(r'[A-Za-z0-9]+', raw_reg)
                                best_match = raw_reg
                                for m in matches:
                                    if any(ch.isdigit() for ch in m):
                                        if len(m) >= 2:
                                            best_match = m
                                            break
                                        
                                all_students.append({
                                    'REG_NO': best_match,
                                    'NAME': name_val,
                                    'DEPT': class_key
                                })
    
    # 3. Dummy PDF reading for timetable (real extraction logic depends on PDF format)
    # reader = pypdf.PdfReader(timetable_file)
    # text = reader.pages[0].extract_text()
    
    # --- Extract subcodes from PDF or Excel ---
    dept_subcodes = {}
    if timetable_file:
        try:
            # Standardise user input exam_date
            try:
                exam_date_std = pd.to_datetime(exam_date, dayfirst=True).strftime('%d-%m-%Y')
            except:
                exam_date_std = exam_date
                
            if hasattr(timetable_file, 'seek'):
                timetable_file.seek(0)
                
            filename = timetable_file.name.lower() if hasattr(timetable_file, 'name') else ""
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                df_tt = pd.read_excel(timetable_file, header=None)
                header_idx = -1
                for idx, row_tt in df_tt.iterrows():
                    row_str = ' '.join([str(x).upper() for x in row_tt.values])
                    if 'DATE' in row_str and 'SESSION' in row_str and 'DEPARTMENT' in row_str:
                        header_idx = idx
                        break
                        
                if header_idx != -1:
                    df_tt.columns = df_tt.iloc[header_idx]
                    df_tt = df_tt[header_idx+1:]
                    
                    c_date = ''
                    c_sess = ''
                    
                    for idx, row_tt in df_tt.iterrows():
                        d = row_tt.get('Date', '')
                        if pd.notnull(d) and str(d).strip() != '':
                            try:
                                c_date = pd.to_datetime(d, dayfirst=True).strftime('%d-%m-%Y')
                            except:
                                c_date = str(d).strip()
                            
                        s = str(row_tt.get('Session', '')).strip()
                        if s != 'nan' and s != '': c_sess = s
                            
                        if c_date == exam_date_std and c_sess.strip().upper() == exam_session.strip().upper():
                            yr = str(row_tt.get('Year', '')).strip().upper()
                            dpt = str(row_tt.get('Department', '')).strip().upper()
                            sc = ''
                            for col_name in df_tt.columns:
                                if isinstance(col_name, str) and ('CODE' in col_name.upper() or 'SUB' in col_name.upper() or 'PAPER' in col_name.upper()):
                                    sc = str(row_tt.get(col_name, '')).strip()
                                    break
                            
                            y_strs = []
                            for y_part in yr.split('/'):
                                y_part = y_part.strip()
                                if y_part == 'I': y_strs.append('I-YEAR')
                                elif y_part == 'II': y_strs.append('II-YEAR')
                                elif y_part == 'III' or y_part == 'OG': y_strs.append('III-YEAR')
                            
                            dept_list = []
                            import re
                            clean_dpt = re.sub(r'[^A-Z]', '', dpt)
                            if 'ALL' in dpt:
                                dept_list = ['CS', 'AI&DS', 'B.COM', 'B A TAMIL', 'CHE', 'BCA', 'BBA', 'MIB']
                            else:
                                dpt_parts = [x.strip() for x in dpt.split(',')]
                                if 'CS' in clean_dpt or 'COMPUTERSCIENCE' in clean_dpt: dept_list.append('CS')
                                if 'AIDS' in clean_dpt or 'ARTIFICIALINTELLIGENCE' in clean_dpt: dept_list.append('AI&DS')
                                if 'BCOM' in clean_dpt or 'COMMERCE' in clean_dpt or 'COM' in clean_dpt: dept_list.append('B.COM')
                                if 'TAM' in clean_dpt or 'TA' in clean_dpt or 'T' in dpt_parts: dept_list.append('B A TAMIL')
                                if 'CHE' in clean_dpt or 'CH' in dpt_parts or 'CHEMISTRY' in clean_dpt: dept_list.append('CHE')
                                if 'BCA' in clean_dpt: dept_list.append('BCA')
                                if 'BBA' in clean_dpt: dept_list.append('BBA')
                                if 'MIB' in clean_dpt: dept_list.append('MIB')
                            
                            for dept_name in dept_list:
                                for y_str in y_strs:
                                    key = f'{dept_name} - {y_str}'
                                    if key not in dept_subcodes or 'TA' in sc:
                                        dept_subcodes[key] = sc
            else:
                reader = pypdf.PdfReader(timetable_file)
                search_str = (exam_date + exam_session).upper().replace(' ', '')
                current_degree = ''
                current_sem = -1
                for page in reader.pages:
                    text = page.extract_text()
                    if not text: continue
                    for line in text.split('\n'):
                        if line.startswith('DEGREE NAME:'):
                            current_degree = line.split('DEGREE NAME:')[1].strip().upper()
                        elif 'SEMESTER :' in line:
                            sem_str = line.split('SEMESTER')[0].strip()
                            if sem_str.isdigit():
                                current_sem = int(sem_str)
                        
                        if search_str in line.replace(' ', '').upper():
                            sub_code = line.split(' ')[0].strip()
                            year = ''
                            if current_sem in [1, 2]: year = '1-YEAR'
                            elif current_sem in [3, 4]: year = '2-YEAR'
                            elif current_sem in [5, 6]: year = '3-YEAR'
                            
                            sheets = []
                            if 'COMPUTER SCIENCE' in current_degree or 'CS' in current_degree: sheets.append('CS')
                            if 'ARTIFICIAL INTELLIGENCE' in current_degree or 'DATA SCIENCE' in current_degree or 'AI&DS' in current_degree: sheets.append('AI&DS')
                            if 'COMMERCE' in current_degree or 'B.COM' in current_degree: sheets.append('B.COM')
                            if 'TAMIL' in current_degree: sheets.append('B A TAMIL')
                            if 'CHEMISTRY' in current_degree or 'CHE' in current_degree: sheets.append('CHE')
                            if 'COMPUTER APPLICATIONS' in current_degree or 'BCA' in current_degree: sheets.append('BCA')
                            if 'FOUNDATION' in current_degree:
                                sheets = ['CS', 'AI&DS', 'B.COM', 'B A TAMIL', 'CHE', 'BCA', 'BBA', 'MIB', 'Sheet3']
                                
                            for s_name in sheets:
                                key = f'{s_name} - {year}'
                                if key not in dept_subcodes or 'TA' in sub_code: 
                                    dept_subcodes[key] = sub_code
        except:
            pass
    # ---------------------------------
    
    # Append SUBCODE to students based on DEPT
    for s in all_students:
        import re
        d_clean = re.sub(r"[^A-Z0-9]", "", s['DEPT'].upper())
        d_clean = d_clean.replace("IIIYEAR", "3YEAR").replace("IIYEAR", "2YEAR").replace("IYEAR", "1YEAR")
        sub = ""
        for k, sc in dept_subcodes.items():
            k_clean = re.sub(r"[^A-Z0-9]", "", k.upper())
            k_clean = k_clean.replace("IIIYEAR", "3YEAR").replace("IIYEAR", "2YEAR").replace("IYEAR", "1YEAR")
            if d_clean == k_clean:
                sub = sc
                break
        s['SUBCODE'] = sub

    # 4. Allocation Logic
    # Group students by Department
    students_by_dept = {}
    for s in all_students:
        dept = s['DEPT']
        if dept not in students_by_dept:
            students_by_dept[dept] = []
        students_by_dept[dept].append(s)
        
    depts = list(students_by_dept.keys())
    
    # Create the output excel in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # We will write a single output sheet simulating the printed layout
        out_rows = []
        out_rows.append([exam_title] + [""] * 11)
        
        for _, hall_row in df_halls.iterrows():
            hall_no = hall_row.get('HALL NO', 'UNKNOWN')
            
            has_rows = 'ROWS' in hall_row and pd.notna(hall_row.get('ROWS'))
            has_cols = 'COLUMNS' in hall_row and pd.notna(hall_row.get('COLUMNS'))
            
            try:
                grid_cols = int(float(hall_row['COLUMNS'])) if has_cols else 6
                grid_rows = int(float(hall_row['ROWS'])) if has_rows else 5
            except (ValueError, TypeError):
                continue
            
            # Ensure total_seats and grid capacity are properly aligned
            if 'TOTAL SEAT' in hall_row and pd.notna(hall_row['TOTAL SEAT']):
                try:
                    total_seats = int(float(hall_row['TOTAL SEAT']))
                    # If physical grid is smaller than requested total seats, expand the grid rows!
                    if total_seats > (grid_rows * grid_cols):
                        grid_rows = total_seats // grid_cols if total_seats % grid_cols == 0 else (total_seats // grid_cols) + 1
                except (ValueError, TypeError):
                    continue # Skip invalid rows like headers
            else:
                total_seats = grid_rows * grid_cols
            
            max_cols = grid_cols * 2
            empty_middle = max_cols - 3
            
            # We have visual columns to fill. Each visual column takes `grid_rows` students.
            import math
            num_blocks = 3
            sub_cols_per_block = math.ceil(grid_cols / 3)
            total_physical_cols = num_blocks * sub_cols_per_block
            
            hall_visual_columns = [[] for _ in range(total_physical_cols)]
            hall_assigned = 0
            
            # Helper to get year string
            def get_year(d_str):
                parts = d_str.split(' - ')
                return parts[1].strip() if len(parts) > 1 else "UNKNOWN"
            
            available_years = sorted(list(set([get_year(d) for d in depts if len(students_by_dept[d]) > 0])))
            
            def is_conflict(cand_d, cand_s, cand_y, other_d, other_s, other_y, check_year=True):
                if check_year and cand_y == other_y: return True
                if cand_d.split(' - ')[0].strip() == other_d.split(' - ')[0].strip(): return True
                if cand_s and other_s and cand_s == other_s: return True
                return False

            for physical_idx in range(total_physical_cols):
                if hall_assigned >= total_seats:
                    break
                    
                block_idx = physical_idx // sub_cols_per_block
                sub_col_idx = physical_idx % sub_cols_per_block
                v_col = sub_col_idx * num_blocks + block_idx
                
                if v_col >= grid_cols:
                    continue
                    
                col_students = []
                target_year = available_years[physical_idx % len(available_years)] if available_years else "UNKNOWN"
                
                while len(col_students) < grid_rows and hall_assigned < total_seats:
                    current_years = sorted(list(set([get_year(d) for d in depts if len(students_by_dept[d]) > 0])))
                    if not current_years: break
                    if target_year not in current_years: target_year = current_years[0]
                    
                    r = len(col_students)
                    left_n = None
                    if physical_idx > 0:
                        prev_block = (physical_idx - 1) // sub_cols_per_block
                        prev_sub = (physical_idx - 1) % sub_cols_per_block
                        prev_v_col = prev_sub * num_blocks + prev_block
                        if prev_v_col < grid_cols and len(hall_visual_columns[prev_v_col]) > r:
                            left_n = hall_visual_columns[prev_v_col][r]
                            
                    front_n = col_students[-1] if r > 0 else None
                    
                    chosen_dept = None
                    candidates = [d for d in depts if len(students_by_dept[d]) > 0]
                    
                    best_score = 999
                    for d in candidates:
                        cand_s = students_by_dept[d][0].get('SUBCODE', '')
                        cand_y = get_year(d)
                        is_tgt = (cand_y == target_year)
                        front_c = is_conflict(d, cand_s, cand_y, front_n['DEPT'], front_n.get('SUBCODE',''), get_year(front_n['DEPT']), check_year=False) if front_n else False
                        left_c = is_conflict(d, cand_s, cand_y, left_n['DEPT'], left_n.get('SUBCODE',''), get_year(left_n['DEPT']), check_year=True) if left_n else False
                        
                        score = 5
                        if is_tgt and not front_c and not left_c: score = 0
                        elif is_tgt and not left_c: score = 1
                        elif not is_tgt and not front_c and not left_c: score = 2
                        elif not is_tgt and not left_c: score = 3
                        elif is_tgt: score = 4
                        
                        if score < best_score:
                            best_score = score
                            chosen_dept = d
                        if best_score == 0: break
                    
                    if chosen_dept is None: break
                    
                    max_allowed_for_hall = total_seats - hall_assigned
                    needed = min(grid_rows - len(col_students), max_allowed_for_hall)
                    take_count = min(needed, len(students_by_dept[chosen_dept]))
                    
                    col_students.extend(students_by_dept[chosen_dept][:take_count])
                    students_by_dept[chosen_dept] = students_by_dept[chosen_dept][take_count:]
                    hall_assigned += take_count
                    
                hall_visual_columns[v_col] = col_students
                
            # Do not strip empty sublists because we need indices to map to physical blocks
                
            if not any(hall_visual_columns):
                continue # Skip empty halls
                
            # Format the block for this hall
            out_rows.append([f"Date & Session: {exam_date} & {exam_session}"] + [""] * empty_middle + [f"HALL NO : {hall_no}", ""])

            
            # Calculate actual assigned total and department names
            total_assigned = sum(len(c) for c in hall_visual_columns)
            
            # Calculate actual assigned total and department counts
            total_assigned = sum(len(c) for c in hall_visual_columns)
            

            
            dept_counts = {}
            for col_students in hall_visual_columns:
                for student in col_students:
                    d = student['DEPT']
                    dept_counts[d] = dept_counts.get(d, 0) + 1
                    
            dept_parts = []
            for d, count in dept_counts.items():
                parts = d.split(" - ")
                dept_name = parts[0]
                roman = ""
                if len(parts) > 1:
                    if parts[1] == "I-YEAR" or parts[1] == "1-YEAR": roman = "I"
                    elif parts[1] == "II-YEAR" or parts[1] == "2-YEAR": roman = "II"
                    elif parts[1] == "III-YEAR" or parts[1] == "3-YEAR": roman = "III"
                
                prefix = f"{roman}-{dept_name}" if roman else dept_name
                
                # Robust lookup for subcode
                subcode = ""
                import re
                d_clean = re.sub(r'[^A-Z0-9]', '', d.upper())
                d_clean = d_clean.replace('IIIYEAR', '3YEAR').replace('IIYEAR', '2YEAR').replace('IYEAR', '1YEAR')
                for k, sc in dept_subcodes.items():
                    k_clean = re.sub(r'[^A-Z0-9]', '', k.upper())
                    k_clean = k_clean.replace('IIIYEAR', '3YEAR').replace('IIYEAR', '2YEAR').replace('IYEAR', '1YEAR')
                    if d_clean == k_clean:
                        subcode = sc
                        break
                        
                if subcode:
                    dept_parts.append(f"{prefix} - {count} ({subcode})")
                else:
                    dept_parts.append(f"{prefix} - {count}")
                    
            dept_str = " & ".join(dept_parts)
            
            out_rows.append([f"DEPT/SUB CODE : {dept_str}"] + [""] * empty_middle + [f"TOTAL : {total_assigned}", ""])
            
            # Build the matrix
            import math
            num_blocks = 3
            sub_cols_per_block = math.ceil(grid_cols / 3)
            actual_max_cols = sub_cols_per_block * num_blocks * 2
            seat_matrix = [['' for _ in range(actual_max_cols)] for _ in range(grid_rows)]
            
            # Map logical columns to grouped physical blocks
            for v_col_idx, col_students in enumerate(hall_visual_columns):
                block_idx = v_col_idx % num_blocks
                sub_col_idx = v_col_idx // num_blocks
                excel_col_start = (block_idx * sub_cols_per_block + sub_col_idx) * 2
                
                for row_idx, student in enumerate(col_students):
                    if row_idx < grid_rows:
                        seat_num = v_col_idx * grid_rows + row_idx + 1
                        
                        # Seat Number on the Left, Reg No on the Right
                        seat_matrix[row_idx][excel_col_start] = seat_num
                        seat_matrix[row_idx][excel_col_start + 1] = student['REG_NO']
            
            # Insert block headers above the matrix
            block_headers = [''] * actual_max_cols
            for b in range(num_blocks):
                start = b * sub_cols_per_block * 2
                if start < actual_max_cols:
                    block_headers[start] = f"COLUMN {b+1:02d}"
            out_rows.append(block_headers)
            
            out_rows.extend(seat_matrix)
            out_rows.append([]) # Blank row
            out_rows.append([]) # Blank row

        if not out_rows:
            raise ValueError("No students could be mapped. Please check if the uploaded Student Details file has 'REG.NO' columns.")

        # At this point, find the global max_cols across all halls to ensure safe DataFrame dimensions
        global_max_cols = max([len(r) for r in out_rows] + [0])
        # Pad all rows to global_max_cols
        padded_out_rows = [r + [''] * (global_max_cols - len(r)) for r in out_rows]
        
        out_df = pd.DataFrame(padded_out_rows)
        out_df.to_excel(writer, index=False, header=False, sheet_name='Seating')
        
        # Format the Excel sheet
        ws = writer.sheets['Seating']
        
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Set column widths
        for col_idx in range(1, global_max_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            if col_idx % 2 != 0:
                ws.column_dimensions[col_letter].width = 10 # SEAT NO columns
            else:
                ws.column_dimensions[col_letter].width = 18 # REGISTER NUMBER columns
                
        # Apply borders and alignments
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=global_max_cols):
            # Check if the row is entirely empty (separator row)
            is_empty_row = all(c.value is None or str(c.value).strip() == "" for c in row)
            
            if not is_empty_row:
                first_val = str(row[0].value) if row[0].value else ""
                is_header_row = (row[0].row == 1) or ("Date & Session" in first_val) or ("DEPT/SUB CODE" in first_val) or ("REGISTER NUMBER" in first_val)
                
                # Determine how many columns this specific hall uses (look for empty cells at the end)
                # But to keep it simple, we draw borders across the entire populated length of this row.
                row_length = 0
                for idx, c in enumerate(row):
                    if c.value is not None and str(c.value).strip() != "":
                        row_length = idx + 1
                
                # If it's a header row, we might want to span it up to the width it should be.
                # Actually, relying on global_max_cols is fine for borders as long as it looks uniform.
                
                for cell in row:
                    # Draw borders only if it's within the populated area or it's a global header
                    if cell.column <= global_max_cols:
                        cell.border = thin_border
                    
                    val_str = str(cell.value) if cell.value else ""
                    
                    # Align center by default
                    if cell.row == 1 or "HALL NO :" in val_str or "TOTAL :" in val_str:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif "Date & Session:" in val_str or "DEPT/SUB CODE :" in val_str:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Make headers bold
                    if is_header_row:
                        if cell.row == 1:
                            cell.font = Font(bold=True, size=14)
                        else:
                            cell.font = Font(bold=True)

        # Merge cells AFTER setting borders
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=global_max_cols):
            first_cell_val = str(row[0].value) if row[0].value else ""
            
            if row[0].row == 1:
                # Top title row
                ws.merge_cells(start_row=row[0].row, start_column=1, end_row=row[0].row, end_column=global_max_cols)
            elif "Date & Session:" in first_cell_val:
                ws.merge_cells(start_row=row[0].row, start_column=1, end_row=row[0].row, end_column=global_max_cols-2)
                ws.merge_cells(start_row=row[0].row, start_column=global_max_cols-1, end_row=row[0].row, end_column=global_max_cols)
            elif "DEPT/SUB CODE :" in first_cell_val:
                ws.merge_cells(start_row=row[0].row, start_column=1, end_row=row[0].row, end_column=global_max_cols-2)
                ws.merge_cells(start_row=row[0].row, start_column=global_max_cols-1, end_row=row[0].row, end_column=global_max_cols)
            elif str(first_cell_val).startswith("COLUMN"):
                # Dynamically merge based on non-empty cells in this row
                col_starts = []
                for c_idx, cell in enumerate(row):
                    if cell.value and str(cell.value).startswith("COLUMN"):
                        col_starts.append(c_idx + 1)
                for i in range(len(col_starts)):
                    start_c = col_starts[i]
                    end_c = col_starts[i+1] - 1 if i + 1 < len(col_starts) else global_max_cols
                    if start_c < end_c:
                        ws.merge_cells(start_row=row[0].row, start_column=start_c, end_row=row[0].row, end_column=end_c)
        
    return output.getvalue()
